from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelColumn:
    key: str
    label: str
    data_type: str = "string"
    preferred_width: float | None = None


@dataclass(frozen=True)
class ExcelWorkbookSpec:
    sheet_name: str
    table_name: str
    columns: tuple[ExcelColumn, ...]
    rows: tuple[Mapping[str, Any], ...]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin", color="B7C9D6")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

MIN_COLUMN_WIDTH = 10.0
MAX_COLUMN_WIDTH = 60.0
DEFAULT_ROW_HEIGHT = 18.0
MAX_ROW_HEIGHT = 96.0


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", name).strip() or "Export"
    return cleaned[:31]


def _safe_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name).strip("_") or "MSA_Export"
    if cleaned[0].isdigit():
        cleaned = f"MSA_{cleaned}"
    return cleaned[:240]


def _literal_text(value: Any) -> str:
    text = str(value)
    # Excel must never interpret source/user strings as executable formulas.
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _excel_value(value: Any, data_type: str) -> Any:
    if value is None:
        return None
    if data_type in {"integer", "decimal", "number"}:
        if isinstance(value, Decimal):
            return float(value)
        return value
    if data_type == "date":
        if isinstance(value, datetime):
            return value.date()
        return value
    if data_type == "datetime":
        return value
    return _literal_text(value)


def _number_format(data_type: str) -> str | None:
    if data_type == "integer":
        return "0"
    if data_type in {"decimal", "number"}:
        return "0.###"
    if data_type == "date":
        return "yyyy-mm-dd"
    if data_type == "datetime":
        return "yyyy-mm-dd hh:mm"
    return None


def _display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _derive_width(column: ExcelColumn, values: Iterable[Any]) -> float:
    if column.preferred_width is not None:
        return max(MIN_COLUMN_WIDTH, min(MAX_COLUMN_WIDTH, float(column.preferred_width)))
    longest = len(column.label)
    for value in values:
        text = _display_text(value)
        longest = max(longest, max((len(line) for line in text.splitlines()), default=0))
    return max(MIN_COLUMN_WIDTH, min(MAX_COLUMN_WIDTH, longest + 2.5))


def _estimated_row_height(values: list[Any], widths: list[float]) -> float:
    max_lines = 1
    for value, width in zip(values, widths):
        text = _display_text(value)
        if not text:
            continue
        logical_lines = 0
        chars_per_line = max(1, int(width - 1))
        for line in text.splitlines() or [text]:
            logical_lines += max(1, math.ceil(len(line) / chars_per_line))
        max_lines = max(max_lines, logical_lines)
    return min(MAX_ROW_HEIGHT, max(DEFAULT_ROW_HEIGHT, 15.0 * max_lines))


def build_excel_workbook(spec: ExcelWorkbookSpec) -> bytes:
    if not spec.columns:
        raise ValueError("Excel export requires at least one column")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_name(spec.sheet_name)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True

    labels = [column.label for column in spec.columns]
    worksheet.append(labels)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = GRID_BORDER
    worksheet.row_dimensions[1].height = 30

    for row in spec.rows:
        values = [_excel_value(row.get(column.key), column.data_type) for column in spec.columns]
        worksheet.append(values)
        row_number = worksheet.max_row
        for column_number, column in enumerate(spec.columns, start=1):
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.alignment = BODY_ALIGNMENT
            cell.border = GRID_BORDER
            number_format = _number_format(column.data_type)
            if number_format:
                cell.number_format = number_format

    widths: list[float] = []
    for index, column in enumerate(spec.columns, start=1):
        values = [row.get(column.key) for row in spec.rows]
        width = _derive_width(column, values)
        widths.append(width)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row_number, row in enumerate(spec.rows, start=2):
        raw_values = [row.get(column.key) for column in spec.columns]
        worksheet.row_dimensions[row_number].height = _estimated_row_height(raw_values, widths)

    if spec.rows:
        table_ref = f"A1:{get_column_letter(len(spec.columns))}{len(spec.rows) + 1}"
        table = Table(displayName=_safe_table_name(spec.table_name), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
