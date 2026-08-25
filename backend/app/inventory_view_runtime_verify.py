from __future__ import annotations

import csv
import io
import json
from decimal import Decimal

from openpyxl import load_workbook

from app.inventory_view_engine import _main_stock_rows, _migration_review_rows
from app.inventory_view_export import inventory_view_export_csv, inventory_view_export_xlsx
from app.shadow_read_api import _query


def main() -> None:
    main_rows = _main_stock_rows(
        q=None,
        mapping_status=None,
        sort_field=None,
        sort_dir=None,
        limit=2000,
        offset=0,
    )
    migration_rows = _migration_review_rows(
        q=None,
        mapping_status=None,
        source_classification=None,
        review_reason=None,
        sort_field=None,
        sort_dir=None,
        limit=2000,
        offset=0,
    )

    current_sum = sum((Decimal(str(row.get("current_qty") or 0)) for row in main_rows), Decimal("0"))
    product_count = int(_query("SELECT COUNT(*)::int AS n FROM products")[0]["n"])
    lot_count = int(_query("SELECT COUNT(*)::int AS n FROM product_lots")[0]["n"])
    transaction_count = int(_query("SELECT COUNT(*)::int AS n FROM inventory_transactions")[0]["n"])
    active_mapping_count = int(
        _query("SELECT COUNT(*)::int AS n FROM product_cms_mappings WHERE mapping_status='ACTIVE_MATCH' AND valid_to IS NULL")[0]["n"]
    )
    accepted_price_count = int(
        _query("SELECT COUNT(*)::int AS n FROM product_cms_mappings WHERE accepted_operational_price IS NOT NULL")[0]["n"]
    )

    assert product_count == 670, product_count
    assert lot_count == 799, lot_count
    assert transaction_count == 679, transaction_count
    assert len(main_rows) == 799, len(main_rows)
    assert len(migration_rows) == 823, len(migration_rows)
    assert current_sum == Decimal("72009.000"), current_sum
    assert active_mapping_count == 0, active_mapping_count
    assert accepted_price_count == 0, accepted_price_count

    export_kwargs = dict(
        preset="main-stock",
        fields="local_item_name,expiry_date,current_qty,catalogue_price,cms_code",
        q=None,
        mapping_status=None,
        source_classification=None,
        review_reason=None,
        sort_field="local_item_name",
        sort_dir="asc",
    )

    xlsx_response = inventory_view_export_xlsx(**export_kwargs)
    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["cache-control"] == "no-store"
    assert xlsx_response.headers["x-msa-export-read-only"] == "true"
    assert xlsx_response.headers["x-msa-database-canonical"] == "false"
    assert xlsx_response.headers["x-msa-migration-baseline-accepted"] == "false"
    workbook = load_workbook(io.BytesIO(xlsx_response.body), data_only=False)
    worksheet = workbook["Main Stock"]
    xlsx_headers = [worksheet.cell(row=1, column=index).value for index in range(1, 6)]
    assert xlsx_headers == ["Items", "Expiry Date", "Current Qty", "Catalogue Price", "CMS Code"], xlsx_headers
    assert worksheet.max_row - 1 == 799, worksheet.max_row - 1
    assert worksheet.freeze_panes == "A2"
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].alignment.wrap_text is True
    assert worksheet["A2"].border.left.style == "thin"
    assert worksheet["B2"].number_format == "mmm-yy", worksheet["B2"].number_format
    assert worksheet["C2"].number_format == "0", worksheet["C2"].number_format
    assert worksheet["D2"].number_format == "0.00", worksheet["D2"].number_format
    assert len(worksheet.tables) == 1

    csv_response = inventory_view_export_csv(**export_kwargs)
    csv_text = csv_response.body.decode("utf-8")
    assert csv_text.startswith("\ufeff")
    csv_records = list(csv.reader(io.StringIO(csv_text.removeprefix("\ufeff"))))
    assert csv_records[0] == xlsx_headers, csv_records[0]
    assert len(csv_records) - 1 == 799, len(csv_records) - 1

    result = {
        "mode": "READ_ONLY_INVENTORY_VIEW_RUNTIME_VERIFY",
        "main_stock_projected_rows": len(main_rows),
        "migration_review_projected_rows": len(migration_rows),
        "main_stock_current_qty_sum": str(current_sum),
        "products": product_count,
        "lots": lot_count,
        "inventory_transactions": transaction_count,
        "active_matches": active_mapping_count,
        "accepted_operational_prices": accepted_price_count,
        "excel_export_rows": worksheet.max_row - 1,
        "excel_export_columns": xlsx_headers,
        "excel_export_sheet": worksheet.title,
        "excel_export_sort": "local_item_name:asc",
        "excel_export_freeze_panes": str(worksheet.freeze_panes),
        "excel_expiry_number_format": worksheet["B2"].number_format,
        "excel_qty_number_format": worksheet["C2"].number_format,
        "excel_price_number_format": worksheet["D2"].number_format,
        "csv_compat_rows": len(csv_records) - 1,
        "mutation": False,
        "database_canonical": False,
        "migration_baseline_accepted": False,
    }
    print(json.dumps(result, sort_keys=True))
    print("inventory_view_runtime=pass excel_export=pass excel_formats=pass csv_compat=pass mutation=false database_canonical=false")


if __name__ == "__main__":
    main()
